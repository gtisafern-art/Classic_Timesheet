"""Вкладка «Отчёт Т-12» — macOS стиль"""

import logging
import os
import calendar
from datetime import date, datetime, timedelta

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout,
                                QLabel, QDateEdit, QPushButton, QTextEdit,
                                QComboBox, QMessageBox, QFileDialog, QApplication,
                                QTableWidget, QTableWidgetItem, QHeaderView, QMenu)
from PySide6.QtCore import Qt, QThread, Signal, QDate
from PySide6.QtGui import QColor, QFont, QKeyEvent

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.repositories.t12_repo import T12Repository
from utils.format import format_hours, is_hours_code

logger = logging.getLogger(__name__)

WEEKDAYS = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']


def _install_copy_handler(table: QTableWidget):
    """Ctrl+C копирует все выделенные ячейки"""
    orig = table.keyPressEvent
    def handler(event: QKeyEvent):
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_C:
            selected = table.selectedRanges()
            if selected:
                rows_set = set()
                for rng in selected:
                    for r in range(rng.topRow(), rng.bottomRow() + 1):
                        rows_set.add(r)
                rows = sorted(rows_set)
                cols_set = set()
                for rng in selected:
                    for c in range(rng.leftColumn(), rng.rightColumn() + 1):
                        cols_set.add(c)
                cols = sorted(cols_set)
                lines = []
                for r in rows:
                    cells = []
                    for c in cols:
                        item = table.item(r, c)
                        cells.append(item.text() if item else "")
                    lines.append("\t".join(cells))
                QApplication.clipboard().setText("\n".join(lines))
                return
        orig(event)
    table.keyPressEvent = handler


class T12Loader(QThread):
    finished = Signal(list)
    error_msg = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = None
        self.repo = None
        self.restaurant_id = ""
        self.start_date = date.today()
        self.end_date = date.today()
        self.position_filters = []

    def run(self):
        from backend.db import MSSQLClient
        from backend.repositories.t12_repo import T12Repository
        
        self.db = MSSQLClient()
        if not self.db.connect():
            self.error_msg.emit("Не удалось подключиться к БД")
            return
        self.repo = T12Repository(self.db)
        try:
            data = self.repo.load_data(self.restaurant_id, self.start_date, self.end_date, self.position_filters)
            self.finished.emit(data)
        except Exception as e:
            self.error_msg.emit(str(e))
        finally:
            if self.db:
                self.db.disconnect()


class T12Tab(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.t12_repo = T12Repository(db)
        self.t12_restaurants = []
        self.t12_data = []
        self.current_start = date.today().replace(day=1)
        self.current_end = date.today()
        self.position_id_map = {}
        self._loader = None

        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        # --- Верх ---
        top = QHBoxLayout()
        top.setSpacing(10)

        top.addWidget(QLabel("Ресторан"))
        self.restaurant_combo = QComboBox()
        self.restaurant_combo.setMinimumWidth(250)
        top.addWidget(self.restaurant_combo)

        top.addWidget(QLabel("Период с"))
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("dd.MM.yyyy")
        self.start_date_edit.setDate(QDate(self.current_start.year, self.current_start.month, self.current_start.day))
        self.start_date_edit.setFixedWidth(120)
        top.addWidget(self.start_date_edit)

        top.addWidget(QLabel("по"))
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("dd.MM.yyyy")
        self.end_date_edit.setDate(QDate(self.current_end.year, self.current_end.month, self.current_end.day))
        self.end_date_edit.setFixedWidth(120)
        top.addWidget(self.end_date_edit)

        self.build_btn = QPushButton("Сформировать")
        self.build_btn.clicked.connect(self._load_timesheet)
        top.addWidget(self.build_btn)

        top.addWidget(QLabel("Должность"))
        self.position_btn = QPushButton("Все должности")
        self.position_btn.setMinimumWidth(200)
        self.position_menu = QMenu(self)
        self.position_btn.setMenu(self.position_menu)
        top.addWidget(self.position_btn)

        top.addStretch()
        layout.addLayout(top)

        # --- Инфо ---
        info_row = QHBoxLayout()
        self.period_label = QLabel("")
        self.period_label.setStyleSheet("color: #86868B; font-size: 11px;")
        info_row.addWidget(self.period_label)
        info_row.addStretch()
        legend = QLabel("Часы(число) | Отпуск(ОТ) | Отпуск б/с(бс) | Больничный(Б)")
        legend.setStyleSheet("font-weight: 600; font-size: 11px; color: #86868B;")
        info_row.addWidget(legend)
        layout.addLayout(info_row)

        # --- Таблица ---
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setVisible(True)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setMinimumSectionSize(20)
        self.table.horizontalHeader().setDefaultSectionSize(30)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setMinimumHeight(150)
        self.table.setStyleSheet(
            "font-family: -apple-system, 'Segoe UI', sans-serif; font-size: 10px;"
            "QTableWidget::item { padding: 0px 2px; }"
        )
        layout.addWidget(self.table, 1)
        _install_copy_handler(self.table)

        # --- Низ ---
        bottom = QHBoxLayout()
        bottom.setSpacing(8)

        self.excel_btn = QPushButton("Excel")
        self.excel_btn.clicked.connect(self._save_excel)
        bottom.addWidget(self.excel_btn)

        clip_btn = QPushButton("Копировать")
        clip_btn.setProperty("class", "secondary")
        clip_btn.clicked.connect(self._copy_clipboard)
        bottom.addWidget(clip_btn)

        bottom.addStretch()
        self.status_label = QLabel("Готово")
        self.status_label.setStyleSheet("color: #34C759; font-size: 11px;")
        bottom.addWidget(self.status_label)
        layout.addLayout(bottom)

    def set_restaurants(self, restaurants: list):
        self.t12_restaurants = restaurants
        self.restaurant_combo.clear()
        for r in restaurants:
            self.restaurant_combo.addItem(r.name)
        if restaurants:
            self.restaurant_combo.setCurrentIndex(0)

    def set_positions(self, positions: list):
        self._populate_position_menu(sorted([p.name for p in positions]))

    def _populate_position_menu(self, names: list):
        self.position_menu.clear()
        all_act = self.position_menu.addAction("Выбрать все")
        all_act.triggered.connect(self._select_all_positions)
        none_act = self.position_menu.addAction("Снять все")
        none_act.triggered.connect(self._deselect_all_positions)
        self.position_menu.addSeparator()
        for name in names:
            action = self.position_menu.addAction(name)
            action.setCheckable(True)
            action.setChecked(True)
            action.toggled.connect(self._on_position_toggled)
        self._update_position_text()

    def _get_position_actions(self):
        return [a for a in self.position_menu.actions() if a.isCheckable()]

    def _get_selected_positions(self):
        return [a.text() for a in self._get_position_actions() if a.isChecked()]

    def _on_position_toggled(self):
        self._update_position_text()
        if self.t12_data:
            self._fill_table()

    def _select_all_positions(self):
        for a in self._get_position_actions():
            a.setChecked(True)

    def _deselect_all_positions(self):
        for a in self._get_position_actions():
            a.setChecked(False)

    def _update_position_text(self):
        selected = self._get_selected_positions()
        total = len(self._get_position_actions())
        if total == 0 or len(selected) == total:
            self.position_btn.setText("Все должности")
        elif len(selected) == 0:
            self.position_btn.setText("Не выбрано")
        elif len(selected) <= 2:
            self.position_btn.setText(", ".join(selected))
        else:
            self.position_btn.setText(f"Выбрано: {len(selected)}")

    def _load_timesheet(self):
        if not self.t12_restaurants:
            return
        idx = self.restaurant_combo.currentIndex()
        if idx < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите ресторан")
            return
        restaurant = self.t12_restaurants[idx]

        s = self.start_date_edit.date().toPython()
        e = self.end_date_edit.date().toPython()
        if s > e:
            s, e = e, s
        self.current_start = s
        self.current_end = e

        days = (e - s).days + 1
        self.period_label.setText(f"{s.strftime('%d.%m.%Y')} — {e.strftime('%d.%m.%Y')} ({days} дн.)")

        self.overlay.show_overlay(self)

        self._loader = T12Loader(self)
        self._loader.restaurant_id = restaurant.id
        self._loader.start_date = s
        self._loader.end_date = e
        self._loader.finished.connect(self._on_data_loaded)
        self._loader.error_msg.connect(self._on_error)
        self._loader.start()

    def _on_data_loaded(self, data: list):
        self.overlay.hide_overlay()
        self.t12_data = data
        positions = set()
        for r in data:
            if r.position_name:
                positions.add(r.position_name)
            if r.target_position_name:
                positions.add(r.target_position_name)
        self._populate_position_menu(sorted(positions))
        self._fill_table()
        self.status_label.setText(f"Записей: {len(data)}")
        self.status_label.setStyleSheet("color: #34C759; font-size: 11px;")

    def _on_error(self, msg: str):
        self.overlay.hide_overlay()
        self.status_label.setText("Ошибка")
        self.status_label.setStyleSheet("color: #FF3B30; font-size: 11px;")
        QMessageBox.critical(self, "Ошибка", msg)

    def _fill_table(self):
        try:
            self._fill_table_impl()
        except Exception as e:
            logger.error("T12 fill_table error: %s", e)
            import traceback
            traceback.print_exc()

    def _fill_table_impl(self):
        grouped = self.t12_repo.group_data(self.t12_data, self.current_start)
        selected = set(self._get_selected_positions())
        all_count = len(self._get_position_actions())
        if selected and len(selected) < all_count:
            grouped = [r for r in grouped if r.main_position in selected or r.target_position in selected]
        logger.info("T12 fill_table: %d grouped rows, %d raw records",
                     len(grouped), len(self.t12_data))
        
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.table.horizontalHeader().setVisible(False)

        days_list = []
        d = self.current_start
        while d <= self.current_end:
            days_list.append(d)
            d += timedelta(days=1)

        num_info_cols = 4
        num_total_cols = 6
        total_cols = num_info_cols + len(days_list) + num_total_cols
        total_rows = 2 + len(grouped) + 1  # заголовки + данные + итого

        self.table.setColumnCount(total_cols)
        self.table.setRowCount(total_rows)
        self.table.setHorizontalHeaderLabels([""] * total_cols)
        for i in range(total_cols):
            self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)

        # Заголовки строки 0: названия колонок
        info_headers = ["ФАМИЛИЯ, ИНИЦИАЛЫ", "ДОЛЖНОСТЬ", "ВИД", "ЗАМЕЩ. ДОЛЖНОСТЬ"]
        total_headers = ["ДНИ", "ЧАСЫ", "ОТ", "Б", "бс", "ПРИМ"]
        for i, h in enumerate(info_headers):
            self._set_header(0, i, h)
        for i, dd in enumerate(days_list):
            self._set_header(0, num_info_cols + i, str(dd.day))
        for i, h in enumerate(total_headers):
            self._set_header(0, num_info_cols + len(days_list) + i, h)

        # Заголовки строки 1: дни недели
        for i in range(num_info_cols):
            self._set_header(1, i, "")
        for i, dd in enumerate(days_list):
            item = self._set_header(1, num_info_cols + i, WEEKDAYS[dd.weekday()])
            if dd.weekday() >= 5:
                item.setForeground(QColor("#FF3B30"))
        for i in range(num_total_cols):
            self._set_header(1, num_info_cols + len(days_list) + i, "")

        # Данные
        bold_font = QFont()
        bold_font.setPointSize(9)
        bold_font.setBold(True)

        row = 2
        for rec in grouped:
            fio = rec.employee_name
            if rec.work_type in ('Замещение', 'Совмещение') and rec.employee_department:
                fio = f"{fio} ({rec.employee_department})"
            self._set_cell(row, 0, fio)

            self._set_cell(row, 1, rec.main_position)
            self._set_cell(row, 2, rec.work_type)
            self._set_cell(row, 3, rec.target_position)

            d = self.current_start
            for i, dd in enumerate(days_list):
                day_data = rec.days.get(d, {'code': '', 'hours': 0})
                code = day_data['code']
                item = self._set_cell(row, num_info_cols + i, code)
                if code == 'ОТ':
                    item.setForeground(QColor("#FF9500"))
                    item.setFont(bold_font)
                elif code == 'Б':
                    item.setForeground(QColor("#FF3B30"))
                    item.setFont(bold_font)
                elif code == 'бс':
                    item.setForeground(QColor("#AF52DE"))
                    item.setFont(bold_font)
                elif is_hours_code(code):
                    item.setForeground(QColor("#34C759"))
                    item.setFont(bold_font)
                d += timedelta(days=1)

            base = num_info_cols + len(days_list)
            self._set_cell(row, base, str(rec.total_days))
            self._set_cell(row, base + 1, format_hours(rec.total_hours))
            self._set_cell(row, base + 2, str(rec.vacation_days))
            self._set_cell(row, base + 3, str(rec.sick_days))
            self._set_cell(row, base + 4, str(rec.without_pay_days))
            self._set_cell(row, base + 5, "")
            row += 1

        # Итого
        tbd, td, th, tv, ts, tw = {}, 0, 0.0, 0, 0, 0
        for rec in grouped:
            td += rec.total_days; th += rec.total_hours
            tv += rec.vacation_days; ts += rec.sick_days; tw += rec.without_pay_days
            for dd, dv in rec.days.items():
                tbd[dd] = tbd.get(dd, 0) + (1 if dv['code'].replace(',','').replace('.','').isdigit() else 0)

        self._set_cell(row, 0, "ИТОГО:", bold_font)
        self._set_cell(row, 1, "", bold_font)
        self._set_cell(row, 2, "", bold_font)
        self._set_cell(row, 3, f"({len(grouped)} чел.)", bold_font)

        d = self.current_start
        for i, dd in enumerate(days_list):
            cnt = tbd.get(d, 0)
            self._set_cell(row, num_info_cols + i, str(cnt) if cnt > 0 else "", bold_font)
            d += timedelta(days=1)

        base = num_info_cols + len(days_list)
        self._set_cell(row, base, str(td), bold_font)
        self._set_cell(row, base + 1, format_hours(th), bold_font)
        self._set_cell(row, base + 2, str(tv), bold_font)
        self._set_cell(row, base + 3, str(ts), bold_font)
        self._set_cell(row, base + 4, str(tw), bold_font)
        self._set_cell(row, base + 5, "", bold_font)

        # Ширина колонок — авто-подбор + минимумы
        self.table.resizeColumnsToContents()
        for i in range(total_cols):
            w = self.table.columnWidth(i)
            if i < num_info_cols:
                mins = [160, 120, 60, 110]
                if w < mins[i]:
                    self.table.setColumnWidth(i, mins[i])
            elif i < num_info_cols + len(days_list):
                if w < 26:
                    self.table.setColumnWidth(i, 26)
            else:
                if w < 34:
                    self.table.setColumnWidth(i, 34)
        self.table.setVisible(True)
        self.table.setUpdatesEnabled(True)
        self.table.update()

    def _set_header(self, row, col, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setBackground(QColor("#F0F0F0"))
        f = QFont()
        f.setPointSize(9)
        f.setBold(True)
        item.setFont(f)
        self.table.setItem(row, col, item)
        return item

    def _set_cell(self, row, col, text, font=None):
        item = QTableWidgetItem(text)
        if col <= 3:
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        else:
            item.setTextAlignment(Qt.AlignCenter)
        if font:
            item.setFont(font)
        else:
            f = QFont()
            f.setPointSize(9)
            item.setFont(f)
        self.table.setItem(row, col, item)
        return item

    def _save_excel(self):
        if not self.t12_data:
            QMessageBox.warning(self, "Предупреждение", "Нет данных")
            return
        rest_name = self.restaurant_combo.currentText()
        fn = f"Табель_{rest_name}_{self.current_start.strftime('%d.%m')}-{self.current_end.strftime('%d.%m.%Y')}.xlsx".replace(" ", "_")
        fp, _ = QFileDialog.getSaveFileName(self, "Excel", fn, "Excel (*.xlsx)")
        if not fp:
            return

        try:
            grouped = self.t12_repo.group_data(self.t12_data, self.current_start)
            dl = []
            d = self.current_start
            while d <= self.current_end:
                dl.append(d)
                d += timedelta(days=1)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Т-12"

            hfont = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            hfill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            ha = Alignment(horizontal='center', vertical='center')
            cf = Font(name='Arial', size=9)
            cc = Alignment(horizontal='center')
            cl = Alignment(horizontal='left')
            bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

            tc = 4 + len(dl) + 6
            ws.merge_cells(f'A1:{get_column_letter(tc)}1')
            ws['A1'].value = f"{rest_name} — Т-12 {self.current_start.strftime('%d.%m.%Y')} – {self.current_end.strftime('%d.%m.%Y')}"
            ws['A1'].font = Font(name='Arial', size=12, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            hdrs = ['ФИО', 'Должность', 'Вид', 'Замещ. должность']
            for dd in dl:
                hdrs.append(str(dd.day))
            hdrs += ['Дни', 'Часы', 'ОТ', 'Б', 'бс', 'Прим.']

            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = bdr

            for ci in range(1, 5):
                ws.cell(row=4, column=ci, value='').border = bdr
            for i, dd in enumerate(dl, 1):
                c = ws.cell(row=4, column=4 + i, value=WEEKDAYS[dd.weekday()])
                c.font = Font(name='Arial', size=9, color='FF0000' if dd.weekday() >= 5 else '555555', bold=dd.weekday() >= 5)
                c.fill = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')
                c.alignment = cc; c.border = bdr

            for ri, rec in enumerate(grouped, 5):
                ws.cell(row=ri, column=1, value=rec.employee_name).font = cf
                ws.cell(row=ri, column=1).alignment = cl; ws.cell(row=ri, column=1).border = bdr
                ws.cell(row=ri, column=2, value=rec.main_position).font = cf
                ws.cell(row=ri, column=2).alignment = cl; ws.cell(row=ri, column=2).border = bdr
                ws.cell(row=ri, column=3, value=rec.work_type).font = cf
                ws.cell(row=ri, column=3).alignment = cc; ws.cell(row=ri, column=3).border = bdr
                ws.cell(row=ri, column=4, value=rec.target_position).font = cf
                ws.cell(row=ri, column=4).alignment = cl; ws.cell(row=ri, column=4).border = bdr

                for i, dd in enumerate(dl, 1):
                    dv = rec.days.get(dd, {'code': '', 'hours': 0})
                    c = ws.cell(row=ri, column=4 + i, value=dv['code'])
                    c.font = cf; c.alignment = cc; c.border = bdr
                    if dv['code'] == 'ОТ':
                        c.font = Font(name='Arial', size=9, color='FF9500', bold=True)
                    elif dv['code'] == 'Б':
                        c.font = Font(name='Arial', size=9, color='FF3B30', bold=True)
                    elif dv['code'] == 'бс':
                        c.font = Font(name='Arial', size=9, color='AF52DE', bold=True)
                    elif is_hours_code(dv['code']):
                        c.font = Font(name='Arial', size=9, color='34C759', bold=True)

                co = 4 + len(dl)
                ws.cell(row=ri, column=co + 1, value=rec.total_days).font = cf
                ws.cell(row=ri, column=co + 1).alignment = cc; ws.cell(row=ri, column=co + 1).border = bdr
                ws.cell(row=ri, column=co + 2, value=round(rec.total_hours, 2)).font = cf
                ws.cell(row=ri, column=co + 2).alignment = cc; ws.cell(row=ri, column=co + 2).border = bdr
                ws.cell(row=ri, column=co + 3, value=rec.vacation_days).font = cf
                ws.cell(row=ri, column=co + 3).alignment = cc; ws.cell(row=ri, column=co + 3).border = bdr
                ws.cell(row=ri, column=co + 4, value=rec.sick_days).font = cf
                ws.cell(row=ri, column=co + 4).alignment = cc; ws.cell(row=ri, column=co + 4).border = bdr
                ws.cell(row=ri, column=co + 5, value=rec.without_pay_days).font = cf
                ws.cell(row=ri, column=co + 5).alignment = cc; ws.cell(row=ri, column=co + 5).border = bdr
                ws.cell(row=ri, column=co + 6, value='').border = bdr

            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 25
            for i in range(1, len(dl) + 1):
                ws.column_dimensions[get_column_letter(4 + i)].width = 5.5
            for i, w in enumerate([6, 7, 4, 4, 4, 8], 1):
                ws.column_dimensions[get_column_letter(4 + len(dl) + i)].width = w

            wb.save(fp)
            QMessageBox.information(self, "Готово", f"Сохранено:\n{fp}")
            if QMessageBox.question(self, "", "Открыть файл?") == QMessageBox.Yes:
                os.startfile(fp)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _copy_clipboard(self):
        text = self.text_edit.toPlainText()
        if not text.strip():
            QMessageBox.warning(self, "Предупреждение", "Нет данных")
            return
        from PySide6.QtWidgets import QApplication
        QApplication.clipboard().setText(text)
        self.status_label.setText("Скопировано в буфер")
